# app.py ✅ Standalone — ONLY:
#   1) Filtro por meses y semanas (defaults: últimos 3 meses)
#   2) Exportar Excel — Simulación por Ejecutivo (mes/intervalo actual)
#      ✅ Promedio inteligente: ignora ceros consecutivos SOLO al inicio por ejecutivo
#   3) Metas del mes (solo ejecutivos activos)
#      ✅ Usuario selecciona MES (default = mes actual)
#      ✅ Intervalo para promedio = últimos 3 meses (default)
#   4) Metas por TEAM (Supervisor), por CENTRO (JV/CC2) y GLOBAL
#   5) ✅ Sanity check: ventas diarias necesarias + ventas hechas + gap (considera días laborables + sábados 1/2 + puentes MX)
#      ✅ GAP ahora es SIGNED: gap = meta - ventas_hechas  (así SIEMPRE: ventas + gap = meta)

import os
import unicodedata
from datetime import datetime, date
from io import BytesIO
from typing import TYPE_CHECKING

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter

# -------------------------------
# CONFIG
# -------------------------------
st.set_page_config(
    page_title="Metas Mensuales CC",
    page_icon="📈",
    layout="wide",
)

# -------------------------------
# THEME (READ ONLY — we do NOT force anything)
# -------------------------------
try:
    theme_base = st.get_option("theme.base") or "light"
except Exception:
    theme_base = "light"

IS_DARK = str(theme_base).lower() == "dark"

# -------------------------------
# NEUTRAL, THEME-FRIENDLY CSS (no forced colors)
# -------------------------------
st.markdown(
    """
<style>
header[data-testid="stHeader"]{ background: rgba(0,0,0,0) !important; }
header[data-testid="stHeader"] [data-testid="stToolbar"]{ background: rgba(0,0,0,0) !important; }
header[data-testid="stHeader"] button,
header[data-testid="stHeader"] svg{
  color: var(--text-color) !important;
  fill: var(--text-color) !important;
}

.stApp{
  background-color: var(--background-color) !important;
  background-image:
    radial-gradient(circle at 1px 1px, rgba(127,127,127,0.14) 1px, transparent 0) !important;
  background-size: 18px 18px !important;
  color: var(--text-color) !important;
}
.block-container{ padding-top: 1.2rem; }

section[data-testid="stSidebar"]{
  background: var(--secondary-background-color) !important;
  border-right: 1px solid rgba(127,127,127,0.25) !important;
}
section[data-testid="stSidebar"] *{ color: var(--text-color) !important; }

section[data-testid="stSidebar"] input,
section[data-testid="stSidebar"] textarea{
  background: var(--background-color) !important;
  border: 1px solid rgba(127,127,127,0.28) !important;
  color: var(--text-color) !important;
  border-radius: 10px !important;
}
section[data-testid="stSidebar"] [data-baseweb="select"] > div{
  background: var(--background-color) !important;
  border: 1px solid rgba(127,127,127,0.28) !important;
  border-radius: 10px !important;
}
section[data-testid="stSidebar"] [data-baseweb="tag"]{
  background: rgba(127,127,127,0.25) !important;
  color: var(--text-color) !important;
  border-radius: 999px !important;
  font-weight: 800 !important;
}
</style>
""",
    unsafe_allow_html=True,
)

# -------------------------------
# HELPERS
# -------------------------------
def normalize_name(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.strip().upper()
    s = " ".join(s.split())
    s = unicodedata.normalize("NFKD", s)
    s = "".join([c for c in s if not unicodedata.combining(c)])
    return s


# ✅ FIX (ONLY En Tránsito join): normalize folio keys so '123.0' matches '123'
def normalize_folio_key(x) -> str:
    """
    Make FOLIO join-compatible between:
      - ventas_no_conciliadas (often numeric -> '123.0')
      - programacion_entrega (varchar -> '123')
    """
    if x is None:
        return ""
    s = str(x).strip()
    if not s or s.lower() in ("nan", "none"):
        return ""

    # common pyodbc numeric artifact: '123456.0' -> '123456'
    if s.endswith(".0"):
        head = s[:-2]
        if head.isdigit():
            return head

    # if it's a float-like string '123.00' etc.
    if "." in s:
        left, right = s.split(".", 1)
        if left.isdigit() and right.strip("0") == "":
            return left

    return s


def _to_excel_bytes(df: pd.DataFrame, sheet_name: str) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        for i, col in enumerate(df.columns, 1):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(55, max_len)
    return out.getvalue()


# ✅ ADDED (shared multi-sheet excel writer — used for Metas Agregadas download too)
def _to_excel_bytes_multi(sheets: dict[str, pd.DataFrame]) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sh, dff in sheets.items():
            shn = sh[:31]
            dff.to_excel(writer, sheet_name=shn, index=False)
            ws = writer.sheets[shn]
            for i, col in enumerate(dff.columns, 1):
                max_len = max(dff[col].astype(str).map(len).max(), len(str(col))) + 2
                ws.column_dimensions[get_column_letter(i)].width = min(55, max_len)
    return out.getvalue()


# -------------------------------
# DB (SQL Server via pyodbc)
# -------------------------------
def get_db_cfg():
    if "db" in st.secrets:
        return {
            "server": st.secrets["db"]["server"],
            "database": st.secrets["db"]["database"],
            "username": st.secrets["db"]["username"],
            "password": st.secrets["db"]["password"],
            "driver": st.secrets["db"].get("driver", "ODBC Driver 17 for SQL Server"),
        }
    return {
        "server": os.getenv("DB_SERVER", ""),
        "database": os.getenv("DB_DATABASE", ""),
        "username": os.getenv("DB_USERNAME", ""),
        "password": os.getenv("DB_PASSWORD", ""),
        "driver": os.getenv("DB_DRIVER", "ODBC Driver 17 for SQL Server"),
    }


@st.cache_data(ttl=600, show_spinner=False)
def read_sql(query: str) -> pd.DataFrame:
    import pyodbc

    cfg = get_db_cfg()
    conn_str = (
        f"DRIVER={{{cfg['driver']}}};"
        f"SERVER={cfg['server']};"
        f"DATABASE={cfg['database']};"
        f"UID={cfg['username']};"
        f"PWD={cfg['password']};"
        "TrustServerCertificate=yes;"
    )
    with pyodbc.connect(conn_str) as conn:
        return pd.read_sql(query, conn)


@st.cache_data(ttl=3600, show_spinner=False)
def load_empleados() -> pd.DataFrame:
    q = r"""
    SELECT
      Tienda AS Centro,
      [Nombre Completo] AS Nombre,
      [Jefe Inmediato],
      Estatus,
      [Fecha Ingreso],
      [Fecha Baja],
      [Canal de Venta],
      Operacion,
      [Tipo Tienda]
    FROM reporte_empleado('EMPRESA_MAESTRA',1,'','') AS e
    WHERE
      [Canal de Venta] IN ('ATT', 'IZZI')
      AND [Operacion] IN ('CONTACT CENTER')
      AND [Tipo Tienda] IN ('VIRTUAL')
      AND (
        Estatus = 'ACTIVO'
        OR (
          Estatus = 'BAJA'
          AND [Fecha Baja] >= DATEADD(MONTH, -1, DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1))
          AND [Fecha Baja] <  DATEADD(MONTH,  1, DATEFROMPARTS(YEAR(GETDATE()), MONTH(GETDATE()), 1))
        )
      )
    """
    df = read_sql(q)
    df["Nombre"] = df["Nombre"].astype(str).str.strip()
    df["Jefe Inmediato"] = df["Jefe Inmediato"].astype(str).str.strip()
    df["Estatus"] = df["Estatus"].astype(str).str.strip()
    df["Fecha Ingreso"] = pd.to_datetime(df["Fecha Ingreso"], errors="coerce")
    df["Fecha Baja"] = pd.to_datetime(df["Fecha Baja"], errors="coerce")

    df["Supervisor"] = df["Jefe Inmediato"].replace({None: "", "None": ""}).astype(str).str.strip()
    df["Supervisor"] = df["Supervisor"].replace({"": "BAJA"})
    return df


def build_ventas_query(start_yyyymmdd: str, end_yyyymmdd: str) -> str:
    return f"""
    SELECT
      FOLIO,
      [PTO. DE VENTA] AS CENTRO,
      [OPERACION PDV],
      [ESTATUS],
      [EJECUTIVO],
      [FECHA DE CAPTURA],
      [PLAN],
      [RENTA SIN IMPUESTOS],
      [PRECIO],
      [SUBREGION]
    FROM reporte_ventas_no_conciliadas('EMPRESA_MAESTRA', 4, '{start_yyyymmdd}', '{end_yyyymmdd}', 1, '19000101', '20990101')
    WHERE
      [OPERACION PDV] = 'CONTACT CENTER'
      AND [PTO. DE VENTA] LIKE 'EXP ATT C CENTER%'
    """


# ✅ Pull "En Tránsito" by FOLIO exactly like Transito Global (Detalle General)
# ✅ FIX: Accent-insensitive WHERE + COALESCE key (FOLIO/Venta)
def build_transito_query(start_yyyymmdd: str, end_yyyymmdd: str) -> str:
    return f"""
    SELECT
      COALESCE(
        NULLIF(LTRIM(RTRIM(CONVERT(varchar(60), [Folio]))), ''),
        NULLIF(LTRIM(RTRIM(CONVERT(varchar(60), [Venta]))), '')
      ) AS FOLIO,
      [Estatus] AS ESTATUS,
      [Venta]   AS VENTA
    FROM reporte_programacion_entrega('empresa_maestra', 4, '{start_yyyymmdd}', '{end_yyyymmdd}')
    WHERE
      [Tienda solicita] LIKE 'EXP ATT C CENTER%'
      AND (
        [Estatus] COLLATE Latin1_General_CI_AI IN (
          'En entrega',
          'Canc Error',
          'Entregado',
          'En preparacion',
          'En preparación',
          'Back Office',
          'Solicitado'
        )
        OR [Estatus] COLLATE Latin1_General_CI_AI LIKE '%entrega%'
        OR [Estatus] COLLATE Latin1_General_CI_AI LIKE '%prepar%'
        OR [Estatus] COLLATE Latin1_General_CI_AI LIKE '%back office%'
        OR [Estatus] COLLATE Latin1_General_CI_AI LIKE '%solicit%'
        OR [Estatus] COLLATE Latin1_General_CI_AI LIKE '%entregad%'
      );
    """


@st.cache_data(ttl=600, show_spinner=False)
def load_transito_flags(start_yyyymmdd: str, end_yyyymmdd: str) -> pd.DataFrame:
    """
    Returns a DF with: FOLIO, IS_TRANSITO
    Logic matches Transito Global 2.0:
      - En Transito if Estatus in (En entrega, En preparacion, Solicitado, Back Office)
      - OR Estatus == Entregado AND Venta is blank

    ✅ FIX: duplicates per FOLIO -> aggregate with MAX so True is not lost
    ✅ FIX: normalize ESTATUS (accents/case) to match reliably
    ✅ FIX: drop empty folios to avoid merge issues
    ✅ FIX: normalize FOLIO join key to avoid '123.0' vs '123'
    """
    q = build_transito_query(start_yyyymmdd, end_yyyymmdd)
    df = read_sql(q)

    if df.empty:
        return pd.DataFrame(columns=["FOLIO", "IS_TRANSITO"])

    # ✅ HERE: normalize FOLIO key
    df["FOLIO"] = df["FOLIO"].apply(normalize_folio_key)
    df = df[df["FOLIO"].notna() & (df["FOLIO"] != "")].copy()

    df["ESTATUS"] = df["ESTATUS"].astype(str).str.strip()
    est_norm = df["ESTATUS"].map(normalize_name)

    venta = df["VENTA"] if "VENTA" in df.columns else pd.Series(pd.NA, index=df.index)

    def _venta_vacia(x):
        if x is None:
            return True
        if isinstance(x, float) and np.isnan(x):
            return True
        s = str(x).strip()
        return s == "" or s.lower() == "nan" or s.lower() == "none"

    venta_vacia = venta.map(_venta_vacia)

    is_transito = est_norm.isin(
        {
            "EN ENTREGA",
            "EN PREPARACION",
            "SOLICITADO",
            "BACK OFFICE",
        }
    ) | (est_norm.eq("ENTREGADO") & venta_vacia)

    out = df[["FOLIO"]].copy()
    out["IS_TRANSITO"] = is_transito.astype(int)

    out = out.groupby("FOLIO", as_index=False)["IS_TRANSITO"].max()
    out["IS_TRANSITO"] = out["IS_TRANSITO"].astype(bool)
    return out


# ✅ NEW (Sanity Check): En tránsito EXACTLY like Transito Global 2.0 "Detalle general programadas"
def build_programadas_query(start_yyyymmdd: str, end_yyyymmdd: str) -> str:
    return f"""
    SELECT
      [Vendedor]       AS VENDEDOR,
      [Estatus]        AS ESTATUS,
      [Venta]          AS VENTA
    FROM reporte_programacion_entrega('empresa_maestra', 4, '{start_yyyymmdd}', '{end_yyyymmdd}')
    WHERE
      [Tienda solicita] LIKE 'EXP ATT C CENTER%'
      AND [Estatus] IN ('En entrega','Canc Error','Entregado','En preparacion','Back Office','Solicitado');
    """


@st.cache_data(ttl=600, show_spinner=False)
def load_programadas_split_by_exec(start_yyyymmdd: str, end_yyyymmdd: str) -> pd.DataFrame:
    """
    Returns per VENDEDOR:
      - hechas_mes (Status == 'Entregado')
      - transito_mes (Status == 'En Transito')
      - total_mes (hechas_mes + transito_mes)
    Logic matches Transito Global 2.0:
      Status = 'En Transito' if Estatus in (En entrega, En preparacion, Solicitado, Back Office)
              OR (Estatus == 'Entregado' and Venta blank)
      Else Status = 'Entregado'
    IMPORTANT: Canc Error is excluded from totals (like flag_Programada in Detalle General).
    """
    q = build_programadas_query(start_yyyymmdd, end_yyyymmdd)
    df = read_sql(q)

    if df.empty:
        return pd.DataFrame(columns=["VENDEDOR", "hechas_mes", "transito_mes", "total_mes", "VEN_NORM"])

    df["VENDEDOR"] = df["VENDEDOR"].astype(str).str.strip()
    df["ESTATUS"] = df["ESTATUS"].astype(str).str.strip()
    est_norm = df["ESTATUS"].map(normalize_name)

    venta = df["VENTA"] if "VENTA" in df.columns else pd.Series(pd.NA, index=df.index)

    def _venta_vacia(x):
        if x is None:
            return True
        if isinstance(x, float) and np.isnan(x):
            return True
        s = str(x).strip()
        return s == "" or s.lower() == "nan" or s.lower() == "none"

    venta_vacia = venta.map(_venta_vacia)

    is_transito = est_norm.isin({"EN ENTREGA", "EN PREPARACION", "SOLICITADO", "BACK OFFICE"}) | (
        est_norm.eq("ENTREGADO") & venta_vacia
    )

    # Exclude Canc Error from totals (Detalle General flag_Programada = Estatus != 'Canc Error')
    is_canc = est_norm.eq("CANC ERROR")
    df = df[~is_canc].copy()
    is_transito = is_transito[~is_canc]

    out = df[["VENDEDOR"]].copy()
    out["transito_mes"] = is_transito.astype(int)
    out["hechas_mes"] = (~is_transito).astype(int)

    g = (
        out.groupby("VENDEDOR", as_index=False)
        .agg(
            hechas_mes=("hechas_mes", "sum"),
            transito_mes=("transito_mes", "sum"),
        )
        .copy()
    )
    g["hechas_mes"] = pd.to_numeric(g["hechas_mes"], errors="coerce").fillna(0).astype(int)
    g["transito_mes"] = pd.to_numeric(g["transito_mes"], errors="coerce").fillna(0).astype(int)
    g["total_mes"] = (g["hechas_mes"] + g["transito_mes"]).astype(int)

    g["VEN_NORM"] = g["VENDEDOR"].map(normalize_name)
    return g


@st.cache_data(ttl=600, show_spinner=False)
def load_ventas(start_yyyymmdd: str, end_yyyymmdd: str) -> pd.DataFrame:
    q = build_ventas_query(start_yyyymmdd, end_yyyymmdd)
    df = read_sql(q)

    df["EJECUTIVO"] = df["EJECUTIVO"].astype(str).str.strip()
    df["CENTRO"] = df["CENTRO"].astype(str).str.strip()
    df["PLAN"] = df["PLAN"].astype(str).str.strip()

    # ✅ HERE: normalize FOLIO key (ONLY for En Tránsito join correctness)
    df["FOLIO"] = df["FOLIO"].apply(normalize_folio_key)

    def fix_centro(c: str) -> str:
        c_up = str(c).upper()
        if "JUAREZ" in c_up:
            return "EXP ATT C CENTER JUAREZ"
        if "CENTER 2" in c_up:
            return "EXP ATT C CENTER 2"
        return c

    df["CENTRO"] = df["CENTRO"].apply(fix_centro)
    df["CentroKey"] = np.where(df["CENTRO"].str.upper().str.contains("JUAREZ", na=False), "JV", "CC2")

    df["EJECUTIVO"] = df["EJECUTIVO"].replace(
        {
            "CESAR JAHACIEL ALONSO GARCIAA": "CESAR JAHACIEL ALONSO GARCIA",
            "VICTOR BETANZO FUENTES": "VICTOR BETANZOS FUENTES",
        }
    )

    df["FECHA DE CAPTURA"] = pd.to_datetime(df["FECHA DE CAPTURA"], errors="coerce")

    # ✅ Fallback flag (will be overridden ONLY for the selected month in Sanity Check)
    df["ESTATUS"] = df["ESTATUS"].astype(str).str.strip()
    est_norm = df["ESTATUS"].map(normalize_name)

    TRANSITO_KEYS = {
        "EN ENTREGA",
        "EN PREPARACION",
        "BACK OFFICE",
        "SOLICITADO",
        "EN TRANSITO",
    }

    is_transito_by_status = (
        est_norm.isin(TRANSITO_KEYS)
        | est_norm.str.contains(r"\bTRANSITO\b", na=False)
        | est_norm.str.contains(r"\bEN ENTREGA\b", na=False)
        | est_norm.str.contains(r"\bEN PREPARACION\b", na=False)
        | est_norm.str.contains(r"\bBACK OFFICE\b", na=False)
        | est_norm.str.contains(r"\bSOLICITADO\b", na=False)
    )

    def _is_blank_series(x: pd.Series) -> pd.Series:
        s = x.copy()
        s = s.replace({None: np.nan, "None": np.nan})
        s = s.astype(str).str.strip()
        return s.isna() | s.eq("") | s.str.lower().eq("nan")

    blank_plan = _is_blank_series(df["PLAN"]) if "PLAN" in df.columns else pd.Series(False, index=df.index)
    blank_precio = _is_blank_series(df["PRECIO"]) if "PRECIO" in df.columns else pd.Series(False, index=df.index)
    blank_renta = _is_blank_series(df["RENTA SIN IMPUESTOS"]) if "RENTA SIN IMPUESTOS" in df.columns else pd.Series(False, index=df.index)

    venta_vacia = blank_plan & blank_precio & blank_renta
    is_entregado = est_norm.str.contains(r"\bENTREGAD", na=False)
    is_transito_entregado_sin_venta = is_entregado & venta_vacia

    df["IS_TRANSITO"] = (is_transito_by_status | is_transito_entregado_sin_venta).astype(bool)

    return df


def add_supervisor_join(ventas_df: pd.DataFrame, empleados_df: pd.DataFrame) -> pd.DataFrame:
    emp = empleados_df[["Nombre", "Supervisor"]].copy()
    emp["Nombre"] = emp["Nombre"].astype(str).str.strip()
    emp["Nombre_norm"] = emp["Nombre"].map(normalize_name)

    v = ventas_df.copy()
    v["EJ_NORM"] = v["EJECUTIVO"].astype(str).map(normalize_name)

    out = v.merge(emp[["Nombre_norm", "Supervisor"]], left_on="EJ_NORM", right_on="Nombre_norm", how="left")
    out["Supervisor"] = out["Supervisor"].fillna("BAJA")
    out.drop(columns=["Nombre_norm"], inplace=True, errors="ignore")
    return out


# -------------------------------
# SIDEBAR — Parameters + Data Load
# -------------------------------
st.sidebar.header("⚙️ Parámetros")

if "last_refresh" not in st.session_state:
    st.session_state["last_refresh"] = None

btn_cols = st.sidebar.columns([1, 1])
with btn_cols[0]:
    if st.button("🔄 Actualizar datos", use_container_width=True):
        st.cache_data.clear()
        st.session_state["last_refresh"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        st.rerun()
with btn_cols[1]:
    st.caption(f"🕒 {st.session_state['last_refresh']}" if st.session_state["last_refresh"] else "")

default_start = "20250801"
default_start_dt = datetime.strptime(default_start, "%Y%m%d").date()
default_end_dt = date.today()

d1, d2 = st.sidebar.columns(2)
start_dt = d1.date_input("Inicio", value=default_start_dt, format="YYYY-MM-DD")
end_dt = d2.date_input("Fin", value=default_end_dt, format="YYYY-MM-DD")

if start_dt > end_dt:
    st.sidebar.error("⚠️ Inicio no puede ser mayor que Fin.")
    st.stop()

start_yyyymmdd = start_dt.strftime("%Y%m%d")
end_yyyymmdd = end_dt.strftime("%Y%m%d")

try:
    with st.spinner("Cargando ventas desde SQL Server…"):
        ventas_raw = load_ventas(start_yyyymmdd, end_yyyymmdd)
        empleados = load_empleados()
        ventas = add_supervisor_join(ventas_raw, empleados)
except Exception as e:
    st.error("❌ No se pudo conectar al SQL Server (timeout / red / VPN / firewall).")
    st.code(str(e))
    st.stop()

if ventas.empty:
    st.error("No hay datos en el rango seleccionado.")
    st.stop()

ventas["T_DT"] = pd.to_datetime(ventas["FECHA DE CAPTURA"], errors="coerce")
ventas = ventas[ventas["T_DT"].notna()].copy()
ventas["T_MonthKey"] = ventas["T_DT"].dt.strftime("%Y-%m")
ventas["T_MonthName"] = ventas["T_DT"].dt.strftime("%B")
ventas["T_MonthLabel"] = ventas["T_MonthKey"] + " (" + ventas["T_MonthName"] + ")"

month_map_all = (
    ventas[["T_MonthKey", "T_MonthLabel"]]
    .dropna()
    .drop_duplicates()
    .sort_values("T_MonthKey")
)

# ✅ NEW: build Fecha Ingreso map (by normalized name) for tenure + nuevos ingresos
emp_ing = empleados[["Nombre", "Fecha Ingreso"]].copy()
emp_ing["Nombre"] = emp_ing["Nombre"].astype(str).str.strip()
emp_ing["EJ_NORM"] = emp_ing["Nombre"].map(normalize_name)
emp_ing["Fecha Ingreso"] = pd.to_datetime(emp_ing["Fecha Ingreso"], errors="coerce")
emp_ing = emp_ing[emp_ing["EJ_NORM"].notna() & (emp_ing["EJ_NORM"] != "")].copy()
# if duplicates, keep earliest ingreso (stable)
ingreso_map_norm = (
    emp_ing.dropna(subset=["Fecha Ingreso"])
    .groupby("EJ_NORM")["Fecha Ingreso"]
    .min()
    .to_dict()
)

# fallback map (from ventas) in case an ejecutivo doesn't exist in empleados
ventas_norm_all = ventas.copy()
ventas_norm_all["EJ_NORM"] = ventas_norm_all["EJECUTIVO"].astype(str).map(normalize_name)
first_dt_ventas_norm = (
    ventas_norm_all.groupby("EJ_NORM")["T_DT"]
    .min()
    .to_dict()
)

def _get_ingreso_dt_for_norm(ej_norm: str):
    v = ingreso_map_norm.get(ej_norm)
    if pd.isna(v) or v is None:
        return first_dt_ventas_norm.get(ej_norm)
    return v

# ✅ ADDED: Global filters (Supervisor + Ejecutivo) applied to ALL tables
st.sidebar.markdown("---")
st.sidebar.subheader("🔎 Filtros (Supervisor / Ejecutivo)")

_sup_all = (
    ventas["Supervisor"].replace({None: "", "None": ""}).astype(str).str.strip().replace({"": "BAJA"})
)
sup_options = sorted(pd.Series(_sup_all.unique()).dropna().tolist())
sup_selected = st.sidebar.multiselect(
    "Supervisor",
    options=sup_options,
    default=[],
    key="flt_supervisor_multi",
)

ventas_flt = ventas.copy()
if sup_selected:
    ventas_flt = ventas_flt[ventas_flt["Supervisor"].isin(sup_selected)].copy()

ej_options = sorted(ventas_flt["EJECUTIVO"].dropna().unique().tolist())
ej_selected = st.sidebar.multiselect(
    "Ejecutivo",
    options=ej_options,
    default=[],
    key="flt_ejecutivo_multi",
)

if ej_selected:
    ventas_flt = ventas_flt[ventas_flt["EJECUTIVO"].isin(ej_selected)].copy()

# ======================================================
# ✅ 1) Filtro por MESES + semanas (defaults: últimos 3 meses)
#     ✅ IMPORTANT: this filter is ONLY used for Simulación por Ejecutivo
# ======================================================
st.markdown("## Filtro por meses y semanas")

if TYPE_CHECKING:
    ventas: pd.DataFrame

df_ctx = ventas_flt.copy()  # ✅ CHANGED (only to apply filters) — same logic

month_map = (
    df_ctx[["T_MonthKey", "T_MonthLabel"]]
    .dropna()
    .drop_duplicates()
    .sort_values("T_MonthKey")
)
m_options = month_map["T_MonthLabel"].tolist()

# ✅ CHANGED: default last 3 months EXCLUDING current month
def _prev_month_keys_excluding_current(n: int = 3) -> list[str]:
    today = date.today()
    cur = pd.Timestamp(today.year, today.month, 1)
    keys = []
    for i in range(1, n + 1):  # 1..n months back => excludes current month
        k = (cur - pd.DateOffset(months=i)).strftime("%Y-%m")
        keys.append(k)
    return keys

desired_keys = _prev_month_keys_excluding_current(3)
key_to_label = dict(zip(month_map["T_MonthKey"], month_map["T_MonthLabel"]))
defaults = [key_to_label[k] for k in desired_keys if k in key_to_label]

# fallback: if not enough months exist in data
if len(defaults) < 3:
    # try: last 3 available excluding current month key if possible
    cur_key = date.today().strftime("%Y-%m")
    avail = month_map.sort_values("T_MonthKey")["T_MonthKey"].tolist()
    avail_no_cur = [k for k in avail if k != cur_key]
    defaults = [key_to_label[k] for k in (avail_no_cur[-3:] if len(avail_no_cur) >= 3 else avail_no_cur)]

prev_m = st.session_state.get("tend_mvw_months_multi", None)
if prev_m is not None and not isinstance(prev_m, list):
    st.session_state["tend_mvw_months_multi"] = defaults

m_sel = st.multiselect(
    "Selecciona uno o más meses",
    options=m_options,
    default=defaults,
    key="tend_mvw_months_multi",
)

df_f = df_ctx.copy()
if m_sel:
    df_f = df_f[df_f["T_MonthLabel"].isin(m_sel)].copy()
else:
    df_f = df_f.iloc[0:0].copy()

month_start = df_f["T_DT"].dt.to_period("M").dt.to_timestamp()
first_wd = month_start.dt.weekday
df_f["T_WeekOfMonth"] = ((df_f["T_DT"].dt.day + first_wd - 1) // 7) + 1
df_f["T_WeekLabel"] = df_f["T_MonthLabel"] + " - Semana " + df_f["T_WeekOfMonth"].astype(int).astype(str)

w_map = (
    df_f[["T_MonthKey", "T_MonthLabel", "T_WeekOfMonth", "T_WeekLabel"]]
    .dropna()
    .drop_duplicates()
    .sort_values(["T_MonthKey", "T_WeekOfMonth"])
)
w_options = w_map["T_WeekLabel"].tolist()

_months_key = tuple(sorted(m_sel)) if m_sel else tuple()
_prev_months_key = st.session_state.get("_tend_prev_months_key", None)
months_changed = (_prev_months_key != _months_key)
st.session_state["_tend_prev_months_key"] = _months_key

prev_weeks = st.session_state.get("tend_mvw_weeks_multi", None)
if months_changed or prev_weeks is None:
    st.session_state["tend_mvw_weeks_multi"] = w_options.copy()
else:
    st.session_state["tend_mvw_weeks_multi"] = [w for w in prev_weeks if w in w_options]

w_sel = st.multiselect(
    "Selecciona Semana(s) del intervalo",
    options=w_options,
    default=w_options,
    key="tend_mvw_weeks_multi",
)

if w_sel:
    df_f = df_f[df_f["T_WeekLabel"].isin(w_sel)].copy()

df_ctx = df_f.copy()
months_selected_keys = sorted(df_ctx["T_MonthKey"].dropna().unique().tolist())

st.markdown("---")

# ======================================================
# ✅ 2) Simulación por Ejecutivo (incluye Supervisor)
#     ✅ Uses df_ctx (interval filter)
# ======================================================
st.markdown("### ⬇️ Exportar Excel — Simulación por Ejecutivo (mes/intervalo actual)")

df_export_base = df_ctx.copy()

if df_export_base.empty:
    st.info("No hay datos para exportar simulación con el filtro actual.")
    st.caption(f"🕒 Render: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    st.stop()

months_selected = sorted(df_export_base["T_MonthKey"].dropna().unique().tolist())
execs = sorted(df_export_base["EJECUTIVO"].dropna().unique().tolist())

sup_map = (
    df_export_base[["EJECUTIVO", "Supervisor"]]
    .dropna(subset=["EJECUTIVO"])
    .copy()
)
sup_map["Supervisor"] = sup_map["Supervisor"].replace({None: "", "None": ""}).astype(str).str.strip()
sup_map["Supervisor"] = sup_map["Supervisor"].replace({"": "BAJA"})
sup_map = (
    sup_map.groupby("EJECUTIVO")["Supervisor"]
    .agg(lambda s: (s.dropna().iloc[0] if len(s.dropna()) else "BAJA"))
    .to_dict()
)

df_month_exec = (
    df_export_base.groupby(["EJECUTIVO", "T_MonthKey"], as_index=False)
    .size()
    .rename(columns={"size": "Ventas_mes"})
)

full_idx = pd.MultiIndex.from_product([execs, months_selected], names=["EJECUTIVO", "T_MonthKey"])
df_month_exec_full = (
    df_month_exec.set_index(["EJECUTIVO", "T_MonthKey"])
    .reindex(full_idx, fill_value=0)
    .reset_index()
)

df_sim = (
    df_month_exec_full.pivot_table(
        index="EJECUTIVO",
        columns="T_MonthKey",
        values="Ventas_mes",
        aggfunc="sum",
        fill_value=0,
    )
    .reset_index()
)

for mk in months_selected:
    if mk not in df_sim.columns:
        df_sim[mk] = 0

month_cols = months_selected[:]
df_sim[month_cols] = df_sim[month_cols].apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)

df_sim["Supervisor"] = df_sim["EJECUTIVO"].map(sup_map).fillna("BAJA")
df_sim["Total ventas"] = df_sim[month_cols].sum(axis=1).astype(int)

vals = df_sim[month_cols].values.astype(int)
n_rows, n_cols = vals.shape
nonzero = vals > 0
has_nonzero = nonzero.any(axis=1)
first_idx = np.where(has_nonzero, nonzero.argmax(axis=1), 0)

suffix_sum = np.cumsum(vals[:, ::-1], axis=1)[:, ::-1]
row_ix = np.arange(n_rows)
sum_active = suffix_sum[row_ix, first_idx]
count_active = (n_cols - first_idx).astype(float)

avg_active = sum_active / count_active
avg_active = np.where(has_nonzero, avg_active, 0.0)

df_sim["Promedio ventas meses"] = avg_active.astype(float)

# ✅ last month of interval = chronologically max (NOT selection order)
last_month_interval = max(month_cols) if month_cols else ""
df_sim["status"] = np.where(df_sim[last_month_interval] > 0, "ACTIVO", "BAJA")

# ✅ Fecha Ingreso (preferred) + fallback to first sale date
df_sim["EJ_NORM"] = df_sim["EJECUTIVO"].astype(str).map(normalize_name)
first_dt_exec = pd.to_datetime(df_sim["EJ_NORM"].map(_get_ingreso_dt_for_norm), errors="coerce")

# ✅ FIX: interval simulates metas for the NEXT month after the last month in the interval (always)
# Example: Oct-Nov-Dec => metas for Jan (ref_day_sim = Jan 1)
if last_month_interval:
    yy2, mm2 = str(last_month_interval).split("-")
    ref_day_sim = pd.Timestamp(year=int(yy2), month=int(mm2), day=1).normalize() + pd.DateOffset(months=1)
else:
    ref_day_sim = pd.Timestamp(date.today().year, date.today().month, 1).normalize()

df_sim["dias_activo_al_1ro"] = (
    (ref_day_sim - first_dt_exec.dt.normalize()).dt.days
    .fillna(0)
    .astype(int)
    .clip(lower=0)
)

prom_sim = df_sim["Promedio ventas meses"].astype(float)
ten_sim = df_sim["dias_activo_al_1ro"].astype(int)

df_sim["meta simulacion"] = np.where(
    ten_sim < 41,
    6,
    np.where(
        prom_sim >= 7,
        np.floor(prom_sim) + 1,
        np.where(
            prom_sim >= 6,
            np.floor(prom_sim) + 2,
            7
        )
    )
).astype(int)

# ✅ NUEVOS INGRESOS (Simulación): people whose Fecha Ingreso is in the TARGET month (next month)
# and started AFTER day 9
target_month_sim = ref_day_sim.strftime("%Y-%m")
nuevos_ingresos = set(
    df_sim.loc[
        (first_dt_exec.dt.strftime("%Y-%m") == str(target_month_sim)) & (first_dt_exec.dt.day > 9),
        "EJECUTIVO",
    ].tolist()
)
nuevos_ingresos_norm = set(normalize_name(x) for x in nuevos_ingresos)

# ✅ NEW: BAJA => meta simulacion = 0
df_sim.loc[df_sim["status"] == "BAJA", "meta simulacion"] = 0

# ✅ NEW: Meta total (includes all ejecutivos, BAJA already forced to 0)
meta_total_all = int(pd.to_numeric(df_sim["meta simulacion"], errors="coerce").fillna(0).sum())
st.markdown(f"**Meta total simulación (todos los ejecutivos): {meta_total_all:,.0f}**")

# ✅ NEW: include dias_activo_al_1ro next to status
df_sim = df_sim[
    ["EJECUTIVO", "Supervisor", "status", "dias_activo_al_1ro"]
    + month_cols
    + ["Total ventas", "Promedio ventas meses", "meta simulacion"]
].copy()

df_sim = df_sim.sort_values(["Total ventas", "Promedio ventas meses"], ascending=False).reset_index(drop=True)

fmt_sim = {c: "{:,.0f}" for c in month_cols + ["dias_activo_al_1ro", "Total ventas", "meta simulacion"]}
fmt_sim.update({"Promedio ventas meses": "{:,.2f}"})


def highlight_rows_sim(row: pd.Series):
    if row.get("status") == "BAJA":
        return ["background-color: #ff1f3d; color: white; font-weight: 900;"] * len(row)
    ej = normalize_name(row.get("EJECUTIVO"))
    if ej in nuevos_ingresos_norm:
        return ["background-color: #ffd166; color: black; font-weight: 900;"] * len(row)
    return [""] * len(row)


st.dataframe(
    df_sim.style.apply(highlight_rows_sim, axis=1).format(fmt_sim),
    hide_index=True,
    width="stretch",
)

excel_bytes_sim = _to_excel_bytes(df_sim, "Simulacion_Ejecutivos")
fname = (
    f"simulacion_ejecutivos_{months_selected[0]}.xlsx"
    if len(months_selected) == 1
    else f"simulacion_ejecutivos_{months_selected[0]}_a_{months_selected[-1]}.xlsx"
)

st.download_button(
    "⬇️ Descargar Excel (Simulación por Ejecutivo)",
    data=excel_bytes_sim,
    file_name=fname,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

# ======================================================
# ✅ 3) Metas del mes (incluye Supervisor + CentroKey)
#     ✅ Uses meta_month_key as the MAIN filter for the rest of the page
# ======================================================
st.markdown("---")
st.markdown("### 🎯 Metas del mes actual (solo ejecutivos activos)")

meta_month_options = month_map_all["T_MonthKey"].tolist()

today_key = date.today().strftime("%Y-%m")
default_meta_key = today_key if today_key in meta_month_options else (meta_month_options[-1] if meta_month_options else today_key)

if "meta_mes_key" in st.session_state and st.session_state["meta_mes_key"] not in meta_month_options:
    st.session_state["meta_mes_key"] = default_meta_key

meta_month_key = st.selectbox(
    "Mes para metas",
    options=meta_month_options,
    index=meta_month_options.index(default_meta_key) if default_meta_key in meta_month_options else 0,
    key="meta_mes_key",
)

# ✅ CHANGED: window = last 3 months EXCLUDING selected month (as requested)
# Example: if meta_month_key == 2026-01 => use 2025-10, 2025-11, 2025-12
meta_idx = meta_month_options.index(meta_month_key) if meta_month_key in meta_month_options else len(meta_month_options) - 1
meta_window_keys = meta_month_options[max(0, meta_idx - 3) : meta_idx]  # 👈 exclude meta_month_key
if len(meta_window_keys) == 0:
    # fallback (keep app alive if very little data)
    meta_window_keys = meta_month_options[max(0, meta_idx - 2) : meta_idx + 1]

df_meta_base = ventas_flt[ventas_flt["T_MonthKey"].isin(meta_window_keys)].copy()
if df_meta_base.empty:
    st.info("No hay datos para el intervalo de 3 meses seleccionado para metas.")
    st.caption(f"🕒 Render: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    st.stop()

sup_map_metas = (
    df_meta_base[["EJECUTIVO", "Supervisor"]]
    .dropna(subset=["EJECUTIVO"])
    .copy()
)
sup_map_metas["Supervisor"] = sup_map_metas["Supervisor"].replace({None: "", "None": ""}).astype(str).str.strip()
sup_map_metas["Supervisor"] = sup_map_metas["Supervisor"].replace({"": "BAJA"})
sup_map_metas = (
    sup_map_metas.groupby("EJECUTIVO")["Supervisor"]
    .agg(lambda s: (s.dropna().iloc[0] if len(s.dropna()) else "BAJA"))
    .to_dict()
)

centro_map = (
    ventas_flt[ventas_flt["T_MonthKey"] == meta_month_key][["EJECUTIVO", "CentroKey"]]
    .dropna(subset=["EJECUTIVO"])
    .copy()
)
if centro_map.empty:
    centro_map = ventas_flt[["EJECUTIVO", "CentroKey"]].dropna(subset=["EJECUTIVO"]).copy()

centro_map = (
    centro_map.groupby("EJECUTIVO")["CentroKey"]
    .agg(lambda s: (s.dropna().iloc[0] if len(s.dropna()) else "CC2"))
    .to_dict()
)

df_meta_me = (
    df_meta_base.groupby(["EJECUTIVO", "T_MonthKey"], as_index=False)
    .size()
    .rename(columns={"size": "Ventas_mes"})
)

meta_execs = sorted(df_meta_base["EJECUTIVO"].dropna().unique().tolist())

full_idx2 = pd.MultiIndex.from_product([meta_execs, meta_window_keys], names=["EJECUTIVO", "T_MonthKey"])
df_meta_full = (
    df_meta_me.set_index(["EJECUTIVO", "T_MonthKey"])
    .reindex(full_idx2, fill_value=0)
    .reset_index()
)

df_metas = (
    df_meta_full.pivot_table(
        index="EJECUTIVO",
        columns="T_MonthKey",
        values="Ventas_mes",
        aggfunc="sum",
        fill_value=0,
    )
    .reset_index()
)

for mk in meta_window_keys:
    if mk not in df_metas.columns:
        df_metas[mk] = 0

df_metas["Supervisor"] = df_metas["EJECUTIVO"].map(sup_map_metas).fillna("BAJA")
df_metas["CentroKey"] = df_metas["EJECUTIVO"].map(centro_map).fillna("CC2")

# status is based on the selected meta month (unchanged intent)
# but the interval columns for prom are meta_window_keys (excluding current month)
df_metas["status"] = np.where(
    ventas_flt[ventas_flt["T_MonthKey"] == meta_month_key].groupby("EJECUTIVO").size().reindex(df_metas["EJECUTIVO"]).fillna(0).values > 0,
    "ACTIVO",
    "BAJA"
)
df_metas = df_metas[df_metas["status"] == "ACTIVO"].copy()

if df_metas.empty:
    st.info("No hay ejecutivos activos en el mes seleccionado para metas.")
    st.caption(f"🕒 Render: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    st.stop()

yy, mm = meta_month_key.split("-")
ref_day = pd.Timestamp(year=int(yy), month=int(mm), day=1).normalize()

# ✅ Fecha Ingreso (preferred) + fallback to first sale date for tenure
df_metas["EJ_NORM"] = df_metas["EJECUTIVO"].astype(str).map(normalize_name)
fd = pd.to_datetime(df_metas["EJ_NORM"].map(_get_ingreso_dt_for_norm), errors="coerce")

tenure_days = (ref_day - fd.dt.normalize()).dt.days
tenure_days = (
    pd.to_numeric(tenure_days, errors="coerce")
    .replace([np.inf, -np.inf], np.nan)
    .fillna(0)
    .astype(int)
    .clip(lower=0)
)
df_metas["dias_activo_al_1ro"] = tenure_days.astype(int)

last3_cols = meta_window_keys[:]
win = df_metas[last3_cols].values.astype(int)
r, c = win.shape

nz = win > 0
has_nz = nz.any(axis=1)
first_i = np.where(has_nz, nz.argmax(axis=1), 0)

suff = np.cumsum(win[:, ::-1], axis=1)[:, ::-1]
ridx = np.arange(r)
ssum = suff[ridx, first_i]
cnt = (c - first_i).astype(float)

avg3 = ssum / cnt
avg3 = np.where(has_nz, avg3, 0.0)
df_metas["prom_ult_3m"] = avg3.astype(float)

prom = df_metas["prom_ult_3m"].astype(float)
ten = df_metas["dias_activo_al_1ro"].astype(int)

meta_mes = np.where(
    ten < 41,
    6,
    np.where(
        prom >= 7,
        np.floor(prom) + 1,
        np.where(
            prom >= 6,
            np.floor(prom) + 2,
            7
        )
    )
).astype(int)

df_metas["meta_mes_actual"] = meta_mes

# ✅ NUEVOS INGRESOS (Metas): Fecha Ingreso in the selected meta_month_key and started AFTER day 9
first_dt_exec2 = pd.to_datetime(df_metas["EJ_NORM"].map(_get_ingreso_dt_for_norm), errors="coerce")
nuevos_ingresos_metas = set(
    df_metas.loc[
        (first_dt_exec2.dt.strftime("%Y-%m") == meta_month_key) & (first_dt_exec2.dt.day > 9),
        "EJECUTIVO",
    ].tolist()
)

# ✅ NEW: normalized set to avoid mismatches (spaces/accents/case)
nuevos_ingresos_metas_norm = set(normalize_name(x) for x in nuevos_ingresos_metas)

df_metas_view = df_metas[
    ["EJECUTIVO", "Supervisor", "CentroKey", "status", "dias_activo_al_1ro"]
    + last3_cols
    + ["prom_ult_3m", "meta_mes_actual"]
].copy()

df_metas_view = df_metas_view.sort_values(
    ["meta_mes_actual", "prom_ult_3m"],
    ascending=[False, False],
).reset_index(drop=True)

fmt_metas = {c: "{:,.0f}" for c in last3_cols + ["dias_activo_al_1ro", "meta_mes_actual"]}
fmt_metas.update({"prom_ult_3m": "{:,.2f}"})


def highlight_metas(row: pd.Series):
    ej = normalize_name(row.get("EJECUTIVO"))
    if ej in nuevos_ingresos_metas_norm:
        return ["background-color: #ffd166; color: black; font-weight: 900;"] * len(row)
    return [""] * len(row)


st.dataframe(
    df_metas_view.style.apply(highlight_metas, axis=1).format(fmt_metas),
    hide_index=True,
    width="stretch",
)

excel_bytes_metas = _to_excel_bytes(df_metas_view, "Metas_Mes")
st.download_button(
    "⬇️ Descargar Excel (Metas mes seleccionado)",
    data=excel_bytes_metas,
    file_name=f"metas_mes_{meta_month_key}_ult3m_{meta_window_keys[0]}_a_{meta_window_keys[-1]}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

# ======================================================
# ✅ 4) Metas por TEAM (Supervisor), por CENTRO y GLOBAL
# ======================================================
st.markdown("---")
st.markdown("### 📌 Metas agregadas: por Team (Supervisor), por Centro y Global")

df_team = (
    df_metas_view.groupby(["Supervisor"], as_index=False)
    .agg(
        ejecutivos=("EJECUTIVO", "nunique"),
        meta_team=("meta_mes_actual", "sum"),
        promedio_meta=("meta_mes_actual", "mean"),
    )
)
df_team["promedio_meta"] = df_team["promedio_meta"].astype(float)
df_team = df_team.sort_values(["meta_team", "ejecutivos"], ascending=[False, False]).reset_index(drop=True)

fmt_team = {"ejecutivos": "{:,.0f}", "meta_team": "{:,.0f}", "promedio_meta": "{:,.2f}"}

st.markdown("#### 🧩 Meta por Team (Supervisor)")
st.dataframe(df_team.style.format(fmt_team), hide_index=True, width="stretch")

df_centro = (
    df_metas_view.groupby(["CentroKey"], as_index=False)
    .agg(
        ejecutivos=("EJECUTIVO", "nunique"),
        meta_centro=("meta_mes_actual", "sum"),
        promedio_meta=("meta_mes_actual", "mean"),
    )
)
df_centro["promedio_meta"] = df_centro["promedio_meta"].astype(float)
df_centro = df_centro.sort_values(["meta_centro", "ejecutivos"], ascending=[False, False]).reset_index(drop=True)

fmt_centro = {"ejecutivos": "{:,.0f}", "meta_centro": "{:,.0f}", "promedio_meta": "{:,.2f}"}

st.markdown("#### 🏢 Meta por Centro (JV / CC2)")
st.dataframe(df_centro.style.format(fmt_centro), hide_index=True, width="stretch")

df_global = pd.DataFrame(
    [{
        "mes": meta_month_key,
        "ejecutivos_activos": int(df_metas_view["EJECUTIVO"].nunique()),
        "meta_global": int(df_metas_view["meta_mes_actual"].sum()),
        "promedio_meta": float(df_metas_view["meta_mes_actual"].mean()) if len(df_metas_view) else 0.0,
    }]
)
fmt_global = {"ejecutivos_activos": "{:,.0f}", "meta_global": "{:,.0f}", "promedio_meta": "{:,.2f}"}

st.markdown("#### 🌎 Meta Global")
st.dataframe(df_global.style.format(fmt_global), hide_index=True, width="stretch")

# ✅ ADDED: One single Excel download for the whole "Metas agregadas" section
metas_agregadas_xlsx = _to_excel_bytes_multi(
    {
        "Meta_por_Team": df_team,
        "Meta_por_Centro": df_centro,
        "Meta_Global": df_global,
    }
)
st.download_button(
    "⬇️ Descargar Excel (Metas agregadas: Team + Centro + Global)",
    data=metas_agregadas_xlsx,
    file_name=f"metas_agregadas_team_centro_global_{meta_month_key}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

# ======================================================
# ✅ 5) Sanity check: ventas diarias necesarias + ventas hechas + gap (día laboral MX, sábados 1/2, sin puentes)
# ======================================================
st.markdown("---")
st.markdown("### 🧪 Sanity check — Ventas diarias necesarias + Ventas hechas + En tránsito + Total + gap (días laborables + sábados 1/2 + sin puentes MX)")


def _nth_weekday_of_month(year: int, month: int, weekday: int, n: int) -> date:
    d = date(year, month, 1)
    shift = (weekday - d.weekday()) % 7
    d = d.replace(day=1 + shift)
    d = d.replace(day=d.day + 7 * (n - 1))
    return d


def mexico_puentes(year: int) -> set[date]:
    return {
        date(year, 1, 1),
        _nth_weekday_of_month(year, 2, 0, 1),
        _nth_weekday_of_month(year, 3, 0, 3),
        date(year, 5, 1),
        date(year, 9, 16),
        _nth_weekday_of_month(year, 11, 0, 3),
        date(year, 12, 25),
    }


def workable_equiv_between(start_d: date, end_d: date) -> float:
    if end_d < start_d:
        return 0.0
    days = pd.date_range(pd.Timestamp(start_d), pd.Timestamp(end_d), freq="D")
    puentes = mexico_puentes(start_d.year)
    total = 0.0
    for dts in days:
        d = dts.date()
        if dts.year != start_d.year:
            puentes = puentes | mexico_puentes(dts.year)

        if d in puentes:
            continue
        wd = dts.weekday()
        if wd <= 4:
            total += 1.0
        elif wd == 5:
            total += 0.5
        else:
            continue
    return float(total)


def month_bounds(ym_key: str) -> tuple[date, date]:
    y, m = ym_key.split("-")
    y = int(y)
    m = int(m)
    start = date(y, m, 1)
    end = (pd.Timestamp(y, m, 1) + pd.offsets.MonthEnd(1)).date()
    return start, end


def workable_days_equiv_month(ym_key: str) -> float:
    start, end = month_bounds(ym_key)
    return workable_equiv_between(start, end)


def workable_days_equiv_elapsed_in_month(ym_key: str, today: date) -> float:
    start, end = month_bounds(ym_key)
    if today < start:
        return 0.0
    cutoff = min(today, end)
    return workable_equiv_between(start, cutoff)


dias_hab_eq_total = workable_days_equiv_month(meta_month_key)
dias_hab_eq_elapsed = workable_days_equiv_elapsed_in_month(meta_month_key, date.today())

# ✅ NEW: días laborables equivalentes RESTANTES desde HOY hasta fin de mes (incluye hoy)
m_ini_tmp, m_fin_tmp = month_bounds(meta_month_key)
today_d = date.today()
start_rem = today_d if today_d > m_ini_tmp else m_ini_tmp
dias_hab_eq_remaining = workable_equiv_between(start_rem, m_fin_tmp) if today_d <= m_fin_tmp else 0.0

# ✅ FIX (ONLY for "desde hoy" rate): if remaining is Saturday-only (0.5), treat as 1.0 day for rate calc
dias_hab_eq_remaining_for_rate = float(dias_hab_eq_remaining)
if 0.0 < dias_hab_eq_remaining_for_rate < 1.0:
    dias_hab_eq_remaining_for_rate = 1.0

if dias_hab_eq_total <= 0:
    st.info("No se pudieron calcular días laborables equivalentes para el mes seleccionado.")
else:
    df_sanity_exec = df_metas_view.copy()

    try:
        m_ini, m_fin = month_bounds(meta_month_key)
        prog_split = load_programadas_split_by_exec(m_ini.strftime("%Y%m%d"), m_fin.strftime("%Y%m%d"))

        df_sanity_exec["EJ_NORM"] = df_sanity_exec["EJECUTIVO"].astype(str).map(normalize_name)
        df_sanity_exec = df_sanity_exec.merge(
            prog_split[["VEN_NORM", "hechas_mes", "transito_mes", "total_mes"]],
            left_on="EJ_NORM",
            right_on="VEN_NORM",
            how="left",
        )
        df_sanity_exec.drop(columns=["VEN_NORM"], inplace=True, errors="ignore")
    except Exception:
        df_sanity_exec["hechas_mes"] = np.nan
        df_sanity_exec["transito_mes"] = np.nan
        df_sanity_exec["total_mes"] = np.nan

    df_sanity_exec["ventas_hechas_mes"] = pd.to_numeric(df_sanity_exec.get("hechas_mes"), errors="coerce").fillna(0).astype(int)
    df_sanity_exec["ventas_en_transito_mes"] = pd.to_numeric(df_sanity_exec.get("transito_mes"), errors="coerce").fillna(0).astype(int)
    df_sanity_exec["total_ventas_hechas_mes"] = pd.to_numeric(df_sanity_exec.get("total_mes"), errors="coerce").fillna(0).astype(int)

    df_sanity_exec.drop(columns=["hechas_mes", "transito_mes", "total_mes"], inplace=True, errors="ignore")

    df_sanity_exec["gap_meta"] = (
        df_sanity_exec["meta_mes_actual"].astype(int)
        - df_sanity_exec["total_ventas_hechas_mes"].astype(int)
    ).astype(int)

    df_sanity_exec["dias_hab_equiv_mes"] = float(dias_hab_eq_total)
    df_sanity_exec["ventas_diarias_necesarias"] = df_sanity_exec["meta_mes_actual"].astype(float) / float(dias_hab_eq_total)

    ratio = 0.0 if dias_hab_eq_total <= 0 else float(dias_hab_eq_elapsed) / float(dias_hab_eq_total)
    ratio = min(max(ratio, 0.0), 1.0)

    df_sanity_exec["esperado_a_hoy"] = np.floor(df_sanity_exec["meta_mes_actual"].astype(float) * ratio + 1e-9).astype(int)
    df_sanity_exec["al_corriente"] = df_sanity_exec["total_ventas_hechas_mes"].astype(int) >= df_sanity_exec["esperado_a_hoy"].astype(int)

    # ✅ NEW: desde HOY — días restantes y ventas diarias necesarias para cumplir meta
    df_sanity_exec["dias_hab_equiv_restantes"] = float(dias_hab_eq_remaining)
    gap_pos = df_sanity_exec["gap_meta"].astype(float).clip(lower=0.0)
    df_sanity_exec["ventas_diarias_necesarias_desde_hoy"] = np.where(
        float(dias_hab_eq_remaining_for_rate) > 0.0,
        gap_pos / float(dias_hab_eq_remaining_for_rate),
        gap_pos,
    )

    df_sanity_exec = df_sanity_exec[
        [
            "EJECUTIVO",
            "Supervisor",
            "CentroKey",
            "dias_activo_al_1ro",
            "ventas_hechas_mes",
            "ventas_en_transito_mes",
            "total_ventas_hechas_mes",
            "meta_mes_actual",
            "gap_meta",
            "dias_hab_equiv_mes",
            "ventas_diarias_necesarias",
            # ✅ NEW
            "dias_hab_equiv_restantes",
            "ventas_diarias_necesarias_desde_hoy",
        ]
    ].copy()

    df_sanity_exec = df_sanity_exec.sort_values(
        ["gap_meta", "ventas_diarias_necesarias_desde_hoy", "ventas_diarias_necesarias"],
        ascending=[False, False, False],
    ).reset_index(drop=True)

    fmt_sanity = {
        "dias_activo_al_1ro": "{:,.0f}",
        "ventas_hechas_mes": "{:,.0f}",
        "ventas_en_transito_mes": "{:,.0f}",
        "total_ventas_hechas_mes": "{:,.0f}",
        "meta_mes_actual": "{:,.0f}",
        "gap_meta": "{:,.0f}",
        "dias_hab_equiv_mes": "{:,.2f}",
        "ventas_diarias_necesarias": "{:,.2f}",
        # ✅ NEW
        "dias_hab_equiv_restantes": "{:,.2f}",
        "ventas_diarias_necesarias_desde_hoy": "{:,.2f}",
    }

    tmp_for_style = df_sanity_exec[["EJECUTIVO", "total_ventas_hechas_mes", "meta_mes_actual"]].copy()
    tmp_for_style["esperado_a_hoy"] = np.floor(tmp_for_style["meta_mes_actual"].astype(float) * ratio + 1e-9).astype(int)
    tmp_for_style["al_corriente"] = tmp_for_style["total_ventas_hechas_mes"].astype(int) >= tmp_for_style["esperado_a_hoy"].astype(int)
    style_corriente = dict(zip(tmp_for_style["EJECUTIVO"], tmp_for_style["al_corriente"].astype(bool)))

    # ✅ ADDED: highlight En tránsito in BLUE (exec table)
    def _blue_transito_exec(row: pd.Series):
        styles = [""] * len(row)
        try:
            v = int(row.get("ventas_en_transito_mes", 0) or 0)
        except Exception:
            v = 0
        if v > 0:
            idx = row.index.get_loc("ventas_en_transito_mes")
            styles[idx] = "background-color: #1e90ff; color: white; font-weight: 900;"
        return styles

    def highlight_gap_dynamic(row: pd.Series):
        styles = [""] * len(row)

        ej = row.get("EJECUTIVO")
        al_corriente = bool(style_corriente.get(ej, True))

        try:
            gap_val = int(row.get("gap_meta", 0) or 0)
        except Exception:
            gap_val = 0

        if (gap_val > 0) and (not al_corriente):
            col_idx = row.index.get_loc("gap_meta")
            styles[col_idx] = "background-color: #ff1f3d; color: white; font-weight: 900;"

        ej_norm = normalize_name(row.get("EJECUTIVO"))
        if ej_norm in nuevos_ingresos_metas_norm:
            for i in range(len(styles)):
                if i != row.index.get_loc("gap_meta"):
                    styles[i] = "background-color: #ffd166; color: black; font-weight: 900;"

        # overlay BLUE for en tránsito (wins on that cell)
        blue = _blue_transito_exec(row)
        for i in range(len(styles)):
            if blue[i]:
                styles[i] = blue[i]

        return styles

    st.markdown("#### 👤 Por ejecutivo (activo)")
    st.dataframe(
        df_sanity_exec.style.apply(highlight_gap_dynamic, axis=1).format(fmt_sanity),
        hide_index=True,
        width="stretch",
    )

    df_sanity_team = (
        df_sanity_exec.groupby(["Supervisor"], as_index=False)
        .agg(
            ejecutivos=("EJECUTIVO", "nunique"),
            ventas_hechas=("ventas_hechas_mes", "sum"),
            ventas_en_transito=("ventas_en_transito_mes", "sum"),
            total_ventas_hechas=("total_ventas_hechas_mes", "sum"),
            meta_team=("meta_mes_actual", "sum"),
            gap_team=("gap_meta", "sum"),
        )
    )
    df_sanity_team["dias_hab_equiv_mes"] = float(dias_hab_eq_total)
    df_sanity_team["ventas_diarias_necesarias"] = df_sanity_team["meta_team"].astype(float) / float(dias_hab_eq_total)
    df_sanity_team["esperado_a_hoy"] = np.floor(df_sanity_team["meta_team"].astype(float) * ratio + 1e-9).astype(int)
    df_sanity_team["al_corriente"] = df_sanity_team["total_ventas_hechas"].astype(int) >= df_sanity_team["esperado_a_hoy"].astype(int)

    # ✅ NEW: desde HOY (team)
    df_sanity_team["dias_hab_equiv_restantes"] = float(dias_hab_eq_remaining)
    gap_team_pos = df_sanity_team["gap_team"].astype(float).clip(lower=0.0)
    df_sanity_team["ventas_diarias_necesarias_desde_hoy"] = np.where(
        float(dias_hab_eq_remaining_for_rate) > 0.0,
        gap_team_pos / float(dias_hab_eq_remaining_for_rate),
        gap_team_pos,
    )

    df_sanity_team = df_sanity_team.sort_values(
        ["gap_team", "ventas_diarias_necesarias_desde_hoy", "ventas_diarias_necesarias"],
        ascending=False
    ).reset_index(drop=True)

    st.markdown("#### 🧩 Por Team (Supervisor)")

    def highlight_gap_team(val, al_corriente):
        try:
            v = int(val or 0)
        except Exception:
            v = 0
        if v > 0 and (not bool(al_corriente)):
            return "background-color: #ff1f3d; color: white; font-weight: 900;"
        return ""

    # ✅ ADDED: blue for ventas_en_transito column (team)
    def _blue_transito_team(row: pd.Series):
        styles = [""] * len(row)
        try:
            v = int(row.get("ventas_en_transito", 0) or 0)
        except Exception:
            v = 0
        if v > 0:
            idx = row.index.get_loc("ventas_en_transito")
            styles[idx] = "background-color: #1e90ff; color: white; font-weight: 900;"
        return styles

    team_gap_styles = [
        highlight_gap_team(g, c) for g, c in zip(df_sanity_team["gap_team"], df_sanity_team["al_corriente"])
    ]

    st.dataframe(
        df_sanity_team.style.format(
            {
                "ejecutivos": "{:,.0f}",
                "ventas_hechas": "{:,.0f}",
                "ventas_en_transito": "{:,.0f}",
                "total_ventas_hechas": "{:,.0f}",
                "meta_team": "{:,.0f}",
                "gap_team": "{:,.0f}",
                "dias_hab_equiv_mes": "{:,.2f}",
                "ventas_diarias_necesarias": "{:,.2f}",
                "esperado_a_hoy": "{:,.0f}",
                # ✅ NEW
                "dias_hab_equiv_restantes": "{:,.2f}",
                "ventas_diarias_necesarias_desde_hoy": "{:,.2f}",
            }
        ).apply(
            lambda r: (
                [""] * r.index.get_loc("gap_team")
                + [team_gap_styles[r.name]]
                + [""] * (len(r) - r.index.get_loc("gap_team") - 1)
            ),
            axis=1,
        ).apply(
            _blue_transito_team,
            axis=1,
        ),
        hide_index=True,
        width="stretch",
    )

    df_sanity_centro = (
        df_sanity_exec.groupby(["CentroKey"], as_index=False)
        .agg(
            ejecutivos=("EJECUTIVO", "nunique"),
            ventas_hechas=("ventas_hechas_mes", "sum"),
            ventas_en_transito=("ventas_en_transito_mes", "sum"),
            total_ventas_hechas=("total_ventas_hechas_mes", "sum"),
            meta_centro=("meta_mes_actual", "sum"),
            gap_centro=("gap_meta", "sum"),
        )
    )
    df_sanity_centro["dias_hab_equiv_mes"] = float(dias_hab_eq_total)
    df_sanity_centro["ventas_diarias_necesarias"] = df_sanity_centro["meta_centro"].astype(float) / float(dias_hab_eq_total)
    df_sanity_centro["esperado_a_hoy"] = np.floor(df_sanity_centro["meta_centro"].astype(float) * ratio + 1e-9).astype(int)
    df_sanity_centro["al_corriente"] = df_sanity_centro["total_ventas_hechas"].astype(int) >= df_sanity_centro["esperado_a_hoy"].astype(int)

    # ✅ NEW: desde HOY (centro)
    df_sanity_centro["dias_hab_equiv_restantes"] = float(dias_hab_eq_remaining)
    gap_centro_pos = df_sanity_centro["gap_centro"].astype(float).clip(lower=0.0)
    df_sanity_centro["ventas_diarias_necesarias_desde_hoy"] = np.where(
        float(dias_hab_eq_remaining_for_rate) > 0.0,
        gap_centro_pos / float(dias_hab_eq_remaining_for_rate),
        gap_centro_pos,
    )

    df_sanity_centro = df_sanity_centro.sort_values(
        ["gap_centro", "ventas_diarias_necesarias_desde_hoy", "ventas_diarias_necesarias"],
        ascending=False
    ).reset_index(drop=True)

    st.markdown("#### 🏢 Por Centro (JV / CC2)")

    # ✅ ADDED: blue for ventas_en_transito column (centro)
    def _blue_transito_centro(row: pd.Series):
        styles = [""] * len(row)
        try:
            v = int(row.get("ventas_en_transito", 0) or 0)
        except Exception:
            v = 0
        if v > 0:
            idx = row.index.get_loc("ventas_en_transito")
            styles[idx] = "background-color: #1e90ff; color: white; font-weight: 900;"
        return styles

    centro_gap_styles = [
        highlight_gap_team(g, c) for g, c in zip(df_sanity_centro["gap_centro"], df_sanity_centro["al_corriente"])
    ]

    st.dataframe(
        df_sanity_centro.style.format(
            {
                "ejecutivos": "{:,.0f}",
                "ventas_hechas": "{:,.0f}",
                "ventas_en_transito": "{:,.0f}",
                "total_ventas_hechas": "{:,.0f}",
                "meta_centro": "{:,.0f}",
                "gap_centro": "{:,.0f}",
                "dias_hab_equiv_mes": "{:,.2f}",
                "ventas_diarias_necesarias": "{:,.2f}",
                "esperado_a_hoy": "{:,.0f}",
                # ✅ NEW
                "dias_hab_equiv_restantes": "{:,.2f}",
                "ventas_diarias_necesarias_desde_hoy": "{:,.2f}",
            }
        ).apply(
            lambda r: (
                [""] * r.index.get_loc("gap_centro")
                + [centro_gap_styles[r.name]]
                + [""] * (len(r) - r.index.get_loc("gap_centro") - 1)
            ),
            axis=1,
        ).apply(
            _blue_transito_centro,
            axis=1,
        ),
        hide_index=True,
        width="stretch",
    )

    df_sanity_global = pd.DataFrame(
        [{
            "mes": meta_month_key,
            "dias_hab_equiv_mes": float(dias_hab_eq_total),
            "ejecutivos_activos": int(df_sanity_exec["EJECUTIVO"].nunique()),
            "ventas_hechas": int(df_sanity_exec["ventas_hechas_mes"].sum()),
            "ventas_en_transito": int(df_sanity_exec["ventas_en_transito_mes"].sum()),
            "total_ventas_hechas": int(df_sanity_exec["total_ventas_hechas_mes"].sum()),
            "meta_global": int(df_sanity_exec["meta_mes_actual"].sum()),
            "gap_global": int(df_sanity_exec["gap_meta"].sum()),
            "ventas_diarias_necesarias": float(df_sanity_exec["meta_mes_actual"].sum()) / float(dias_hab_eq_total),
        }]
    )
    df_sanity_global["esperado_a_hoy"] = int(np.floor(df_sanity_global["meta_global"].astype(float).iloc[0] * ratio + 1e-9))
    df_sanity_global["al_corriente"] = int(df_sanity_global["total_ventas_hechas"].iloc[0]) >= int(df_sanity_global["esperado_a_hoy"].astype(int).iloc[0])

    # ✅ NEW: desde HOY (global)
    df_sanity_global["dias_hab_equiv_restantes"] = float(dias_hab_eq_remaining)
    gap_global_pos = df_sanity_global["gap_global"].astype(float).clip(lower=0.0)
    df_sanity_global["ventas_diarias_necesarias_desde_hoy"] = np.where(
        float(dias_hab_eq_remaining_for_rate) > 0.0,
        gap_global_pos / float(dias_hab_eq_remaining_for_rate),
        gap_global_pos,
    )

    st.markdown("#### 🌎 Global")

    def highlight_gap_global(row: pd.Series):
        styles = [""] * len(row)
        try:
            gap_val = int(row.get("gap_global", 0) or 0)
        except Exception:
            gap_val = 0
        al_corr = bool(row.get("al_corriente", True))
        if (gap_val > 0) and (not al_corr):
            col_idx = row.index.get_loc("gap_global")
            styles[col_idx] = "background-color: #ff1f3d; color: white; font-weight: 900;"
        # ✅ ADDED: blue for ventas_en_transito (global)
        try:
            vt = int(row.get("ventas_en_transito", 0) or 0)
        except Exception:
            vt = 0
        if vt > 0:
            col_idx = row.index.get_loc("ventas_en_transito")
            styles[col_idx] = "background-color: #1e90ff; color: white; font-weight: 900;"
        return styles

    st.dataframe(
        df_sanity_global.style.format(
            {
                "dias_hab_equiv_mes": "{:,.2f}",
                "ejecutivos_activos": "{:,.0f}",
                "ventas_hechas": "{:,.0f}",
                "ventas_en_transito": "{:,.0f}",
                "total_ventas_hechas": "{:,.0f}",
                "meta_global": "{:,.0f}",
                "gap_global": "{:,.0f}",
                "ventas_diarias_necesarias": "{:,.2f}",
                "esperado_a_hoy": "{:,.0f}",
                # ✅ NEW
                "dias_hab_equiv_restantes": "{:,.2f}",
                "ventas_diarias_necesarias_desde_hoy": "{:,.2f}",
            }
        ).apply(highlight_gap_global, axis=1),
        hide_index=True,
        width="stretch",
    )

    df_exec_export = df_sanity_exec.merge(
        tmp_for_style[["EJECUTIVO", "esperado_a_hoy", "al_corriente"]],
        on="EJECUTIVO",
        how="left",
    )

    sanity_xlsx = _to_excel_bytes_multi(
        {
            "Sanity_Ejecutivo": df_exec_export,
            "Sanity_Team": df_sanity_team,
            "Sanity_Centro": df_sanity_centro,
            "Sanity_Global": df_sanity_global,
        }
    )

    st.download_button(
        "⬇️ Descargar Excel (Sanity check)",
        data=sanity_xlsx,
        file_name=f"sanity_check_ventas_gap_diarias_{meta_month_key}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.caption(f"🕒 Render: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
